#!/usr/bin/env python3
"""Generate realistic-but-100%-synthetic sample data files for testing.

Produces examples/samples/ with:
  - sample_committee_config.yaml    — two committees, joint fundraising
  - sample_donors.xlsx              — 45+ donors across every schedule bucket
  - sample_donors.csv               — same data, CSV form
  - sample_donors_givebutter.csv    — same donors re-shaped as a Givebutter export
                                      (exercises the column mapper's platform-specific aliases)
  - sample_disbursements.xlsx       — 17 vendors (exceeds 5-row page → supplement)
  - sample_debts.xlsx               — 2 outstanding debts

ALL data is synthetic. No real person, committee, vendor, or address appears.
"""
from __future__ import annotations
import csv
import random
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

REPO = Path(__file__).resolve().parent.parent.parent
OUT = REPO / "examples" / "samples"
OUT.mkdir(parents=True, exist_ok=True)

# Deterministic output — same seed → same files every build
random.seed(20260415)

# ---------------------------------------------------------------------------
# Synthetic name pools (CLEARLY FAKE — no real people)
# ---------------------------------------------------------------------------
FIRST_NAMES = [
    "Jane", "Alex", "Sam", "Casey", "Jordan", "Taylor", "Morgan", "Avery",
    "Riley", "Cameron", "Dakota", "Drew", "Emerson", "Finley", "Hayden",
    "Kai", "Lane", "Micah", "Parker", "Quinn", "Reese", "Rowan",
    "Sage", "Skylar", "Sydney", "Blake", "Charlie", "Devin", "Elliot",
    "Frankie", "Harper", "Jamie", "Kendall", "Logan", "Marlow", "Nico",
    "Oakley", "Peyton", "River", "Shiloh",
]
LAST_NAMES = [
    "Example", "Placeholder", "Testing", "Sample", "Synthetic", "Fictional",
    "Demo", "Fake", "Imaginary", "Notreal", "Stand-In", "Mock",
]
AZ_CITIES = [
    ("Phoenix", ["85001", "85003", "85004", "85006", "85008", "85012", "85014", "85016"]),
    ("Tempe", ["85281", "85282", "85283", "85284"]),
    ("Mesa", ["85201", "85203", "85205", "85208", "85213"]),
    ("Chandler", ["85224", "85225", "85226", "85286"]),
    ("Gilbert", ["85233", "85234", "85295", "85296", "85297"]),
    ("Scottsdale", ["85251", "85254", "85260"]),
    ("Glendale", ["85301", "85308"]),
]
OOS_LOCATIONS = [
    ("Columbia",  "MO", "65201"),
    ("Herriman",  "UT", "84096"),
    ("Boone",     "NC", "28607"),
    ("Beaverton", "OR", "97007"),
    ("Austin",    "TX", "78701"),
    ("Seattle",   "WA", "98101"),
]


def rand_az_address():
    street = f"{random.randint(100,9999)} E {random.choice(['Main','Oak','Pine','Elm','Cactus','Desert','Palm','Mesa'])} {random.choice(['St','Ave','Rd','Dr','Blvd','Ln'])}"
    city, zips = random.choice(AZ_CITIES)
    return street, city, "AZ", random.choice(zips)


def rand_oos_address():
    street = f"{random.randint(100,9999)} {random.choice(['Main','1st','Elm','Park','Lake'])} {random.choice(['St','Ave','Rd'])}"
    city, state, zp = random.choice(OOS_LOCATIONS)
    return street, city, state, zp


# ---------------------------------------------------------------------------
# Committee config
# ---------------------------------------------------------------------------
COMMITTEE_CONFIG = {
    "committees": [
        {"id": "9999001", "name": "Jane Example Election Committee",
         "candidate": "Jane Example", "treasurer": "Jane Example",
         "address": "123 Main St, Phoenix AZ 85001",
         "office_sought": "Example District 1",
         "office_type": "special_district",
         "starting_balance_q1": 500.00},
        {"id": "9999002", "name": "Alex Example Election Committee",
         "candidate": "Alex Example", "treasurer": "Alex Example",
         "address": "123 Main St, Phoenix AZ 85001",
         "office_sought": "Example District 1",
         "office_type": "special_district",
         "starting_balance_q1": 0.00},
    ],
    "reports": [
        {"committee_id": "9999001", "period": "Q1",
         "date_range": ["2026-01-01", "2026-03-31"], "filing_date": "2026-04-15"},
        {"committee_id": "9999002", "period": "Q1",
         "date_range": ["2026-01-01", "2026-03-31"], "filing_date": "2026-04-15"},
    ],
    "joint_fundraising": [
        {"campaign": "Example Joint Campaign",
         "split": {"9999001": 0.5, "9999002": 0.5}},
    ],
    "signatures": {
        "9999001": {"image": "", "typed_name": "Jane Example"},
        "9999002": {"image": "", "typed_name": "Alex Example"},
    },
}


def write_committee_config():
    out = OUT / "sample_committee_config.yaml"
    header = ("# openfile-az SAMPLE committee config\n"
              "# 100% synthetic — no real committees or people.\n"
              "# Safe to share. Use as a starting point for your own filing.\n\n")
    out.write_text(header + yaml.safe_dump(COMMITTEE_CONFIG, sort_keys=False, width=100))
    print(f"wrote {out}")


# ---------------------------------------------------------------------------
# Donors — exercise every classification path
# ---------------------------------------------------------------------------

def build_donor_rows():
    rows = []
    d0 = date(2026, 1, 15)

    # 3 in-state individuals >$100 (Schedule A(1)(a))
    for i, amt in enumerate([250.00, 500.00, 250.00]):
        street, city, state, zp = rand_az_address()
        rows.append({
            "date": d0 + timedelta(days=i*7),
            "first": random.choice(FIRST_NAMES),
            "last": random.choice(LAST_NAMES),
            "address": street, "city": city, "state": state, "zip": zp,
            "email": f"donor{i+1}@example.test",
            "occupation": random.choice(["Attorney", "Teacher", "Engineer", "Nurse", ""]),
            "employer": random.choice(["Self-employed", "Example University", "Example Corp", ""]),
            "source": "Example Joint Campaign", "amount": amt,
            "kind": "cash", "donor_type": "individual", "pac_id": "",
            "note": "Joint fundraiser; 50/50 split handled by config",
        })

    # 30 in-state individuals ≤$100 (Schedule A(1)(b) aggregate)
    for i in range(30):
        street, city, state, zp = rand_az_address()
        amt = round(random.choice([5, 10, 25, 50, 55, 75, 100]) + random.random()*0.01, 2)
        rows.append({
            "date": d0 + timedelta(days=i),
            "first": random.choice(FIRST_NAMES),
            "last": random.choice(LAST_NAMES),
            "address": street, "city": city, "state": state, "zip": zp,
            "email": f"donor_small_{i:02d}@example.test",
            "occupation": "", "employer": "",
            "source": "Example Joint Campaign", "amount": amt,
            "kind": "cash", "donor_type": "individual", "pac_id": "",
            "note": "",
        })

    # 5 out-of-state individuals (Schedule A(1)(c))
    for i in range(5):
        street, city, state, zp = rand_oos_address()
        amt = round(random.choice([10, 25, 50, 75]) + random.random()*0.01, 2)
        rows.append({
            "date": d0 + timedelta(days=35 + i),
            "first": random.choice(FIRST_NAMES),
            "last": random.choice(LAST_NAMES),
            "address": street, "city": city, "state": state, "zip": zp,
            "email": f"donor_oos_{i+1}@example.test",
            "occupation": "", "employer": "",
            "source": "Example Joint Campaign", "amount": amt,
            "kind": "cash", "donor_type": "individual", "pac_id": "",
            "note": "OOS donor",
        })

    # 2 PAC contributions (Schedule A(1)(e))
    rows.append({
        "date": date(2026, 3, 24), "first": "", "last": "Example Climate PAC",
        "address": "100 Example Ave", "city": "Washington", "state": "DC", "zip": "20001",
        "email": "", "occupation": "", "employer": "",
        "source": "check", "amount": 2000.00,
        "kind": "cash", "donor_type": "pac", "pac_id": "C00000000",
        "note": "Independent expenditure PAC",
    })
    rows.append({
        "date": date(2026, 3, 7), "first": "", "last": "Example State PAC",
        "address": "PO Box 1", "city": "Tucson", "state": "AZ", "zip": "85701",
        "email": "", "occupation": "", "employer": "",
        "source": "check", "amount": 250.00,
        "kind": "cash", "donor_type": "pac", "pac_id": "200000000",
        "note": "Local advocacy PAC",
    })

    # 1 in-kind contribution (Schedule A(5))
    rows.append({
        "date": date(2026, 3, 31), "first": "", "last": "Example Advocacy PAC",
        "address": "200 Example Blvd", "city": "Los Angeles", "state": "CA", "zip": "90029",
        "email": "", "occupation": "", "employer": "",
        "source": "in_kind_services", "amount": 159.70,
        "kind": "in_kind", "donor_type": "pac", "pac_id": "C00000001",
        "note": "March staff / phone / text services",
    })

    return rows


def write_donors(rows):
    # XLSX (our canonical template shape)
    xlsx = OUT / "sample_donors.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Donors"
    cols = list(rows[0].keys())
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = PatternFill("solid", start_color="305496")
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(rows, start=2):
        for i, c in enumerate(cols, start=1):
            v = row[c]
            if isinstance(v, date):
                v = v.strftime("%Y-%m-%d")
            ws.cell(row=r, column=i, value=v)
    for col_idx, w in enumerate([12, 12, 14, 28, 14, 8, 10, 28, 14, 18, 24, 10, 10, 14, 14, 34], start=1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"
    wb.save(xlsx)
    print(f"wrote {xlsx}")

    # CSV (same data, simpler format)
    csv_path = OUT / "sample_donors.csv"
    with open(csv_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for row in rows:
            r = dict(row)
            if isinstance(r["date"], date):
                r["date"] = r["date"].strftime("%Y-%m-%d")
            w.writerow(r)
    print(f"wrote {csv_path}")

    # CSV with Givebutter-style headers (exercises column_mapper aliases)
    gb_path = OUT / "sample_donors_givebutter.csv"
    with open(gb_path, "w", newline="") as f:
        fieldnames = [
            "Campaign Title", "First Name", "Last Name", "Email", "Phone",
            "Address Line 1", "City", "State", "Postal Code", "Country",
            "Amount", "Fee", "Transaction Date (UTC)", "Method", "Internal Note",
        ]
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for row in rows:
            if row["donor_type"] != "individual" or row["kind"] != "cash":
                continue
            d = row["date"]
            d_str = d.strftime("%Y-%m-%d %H:%M:%S") if isinstance(d, date) else str(d)
            w.writerow({
                "Campaign Title": row["source"],
                "First Name": row["first"],
                "Last Name": row["last"],
                "Email": row["email"],
                "Phone": "",
                "Address Line 1": row["address"],
                "City": row["city"],
                "State": row["state"],
                "Postal Code": row["zip"],
                "Country": "US",
                "Amount": f"${row['amount']:.2f}",
                "Fee": "$0.00",
                "Transaction Date (UTC)": d_str,
                "Method": "card",
                "Internal Note": row["note"],
            })
    print(f"wrote {gb_path}")

    # CSV with ActBlue-style headers (43 columns; only individual cash rows
    # since ActBlue doesn't carry PAC or in-kind records).
    ab_path = OUT / "sample_donors_actblue.csv"
    with open(ab_path, "w", newline="") as f:
        fieldnames = [
            "Lineitem ID", "Order Number", "Processor Txn ID", "Amount", "Txn Amount",
            "Stripe Fee Amount", "Fee", "Net Settlement", "Contribution Datetime",
            "Paid At", "Created Datetime", "Payout Datetime", "Transaction Type",
            "Description", "Is Recurring", "Recurrence Number", "Pledged Recurring Duration",
            "Recipient", "Recipient Committee", "Recipient ID", "Recipient Gov ID",
            "Donor First Name", "Donor Last Name", "Donor Address Line 1", "Donor City",
            "Donor State", "Donor ZIP", "Donor Country", "Donor Occupation", "Donor Employer",
            "Donor Email", "Donor Phone", "Employer Address Line 1", "Employer City",
            "Employer State", "Employer Zip", "Employer Country", "Via Mobile",
            "Recurring Amount", "Initial Recurring Contribution Date", "Cancelled Recurring?",
            "Donor U.S. Passport Number", "Payment Method",
        ]
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        line_seq = 1000
        for row in rows:
            if row["donor_type"] != "individual" or row["kind"] != "cash":
                continue
            line_seq += 1
            d = row["date"]
            dt_str = d.strftime("%Y-%m-%d %H:%M:%S UTC") if isinstance(d, date) else str(d)
            d_str = d.strftime("%Y-%m-%d") if isinstance(d, date) else str(d)
            gross = row["amount"]
            # Synthetic processor fee for realism (not representative of actual fees)
            fee = round(gross * 0.04 + 0.30, 2)
            net = round(gross - fee, 2)
            w.writerow({
                "Lineitem ID": f"L{line_seq:09d}",
                "Order Number": f"O{line_seq:09d}",
                "Processor Txn ID": f"ch_{line_seq:024x}",
                "Amount": f"{gross:.2f}",
                "Txn Amount": f"{gross:.2f}",
                "Stripe Fee Amount": f"{fee:.2f}",
                "Fee": f"{fee:.2f}",
                "Net Settlement": f"{net:.2f}",
                "Contribution Datetime": dt_str,
                "Paid At": dt_str,
                "Created Datetime": dt_str,
                "Payout Datetime": d_str,
                "Transaction Type": "contribution",
                "Description": row["source"],
                "Is Recurring": "false",
                "Recurrence Number": "",
                "Pledged Recurring Duration": "",
                "Recipient": "Example Slate",
                "Recipient Committee": "Jane Example Election Committee",
                "Recipient ID": "9999001",
                "Recipient Gov ID": "",
                "Donor First Name": row["first"],
                "Donor Last Name": row["last"],
                "Donor Address Line 1": row["address"],
                "Donor City": row["city"],
                "Donor State": row["state"],
                "Donor ZIP": row["zip"],
                "Donor Country": "US",
                "Donor Occupation": row["occupation"],
                "Donor Employer": row["employer"],
                "Donor Email": row["email"],
                "Donor Phone": "",
                "Employer Address Line 1": "",
                "Employer City": "",
                "Employer State": "",
                "Employer Zip": "",
                "Employer Country": "",
                "Via Mobile": "false",
                "Recurring Amount": "",
                "Initial Recurring Contribution Date": "",
                "Cancelled Recurring?": "",
                "Donor U.S. Passport Number": "",
                "Payment Method": "card",
            })
    print(f"wrote {ab_path}")


# ---------------------------------------------------------------------------
# Disbursements — 17 vendors to exercise 5-row-per-page overflow
# ---------------------------------------------------------------------------
DISBURSEMENTS = [
    ("2026-03-31", "Example Print Vendor A", "100 Print Rd", "Phoenix", "AZ", "85001",
     "Print / mailer", 509.26, "B(1)", "cash", "", "Q1 mailer batch 1"),
    ("2026-03-31", "Example Print Vendor B", "200 Print Rd", "Phoenix", "AZ", "85001",
     "Print / mailer", 1021.86, "B(1)", "cash", "", "Q1 mailer batch 2"),
    ("2026-03-15", "Example Sign Shop", "300 Trade Way", "Austin", "TX", "78701",
     "Door hangers", 67.75, "B(1)", "cash", "", ""),
    ("2026-03-20", "Example Flyer Co", "400 Beach Blvd", "Huntington Beach", "CA", "92601",
     "Door hangers", 100.58, "B(1)", "cash", "", ""),
    ("2026-03-05", "Example Card Printer", "500 Henderson Rd", "Columbus", "OH", "43201",
     "Business cards", 77.64, "B(1)", "cash", "", ""),
    ("2026-03-31", "Example AI Tool", "600 Market St", "San Francisco", "CA", "94101",
     "SaaS — AI subscription", 213.42, "B(1)", "credit", "", ""),
    ("2026-03-31", "Example Routing API", "700 Market St", "San Francisco", "CA", "94101",
     "SaaS — AI API", 116.02, "B(1)", "credit", "", ""),
    ("2026-03-31", "Example Hosting Co", "800 Barranca Ave", "Covina", "CA", "91701",
     "SaaS — hosting", 40.31, "B(1)", "credit", "", ""),
    ("2026-03-31", "Example EU Hosting", "900 Industrie Str", "Berlin", "DE", "10115",
     "SaaS — hosting (overseas)", 6.87, "B(1)", "credit", "", ""),
    ("2026-03-31", "Example Design App", "1000 Design Ln", "Sydney", "AU", "2000",
     "SaaS — design", 8.19, "B(1)", "credit", "", ""),
    ("2026-03-31", "Example Maps API", "1100 Map Ave", "Washington", "DC", "20001",
     "SaaS — mapping", 14.73, "B(1)", "credit", "", ""),
    ("2026-03-31", "Example Office Suite", "1200 Campus Dr", "Mountain View", "CA", "94001",
     "SaaS — email/office", 50.49, "B(1)", "credit", "", ""),
    ("2026-03-31", "Example Telephony API", "1300 Market St", "San Francisco", "CA", "94101",
     "SaaS — telephony", 17.51, "B(1)", "credit", "", ""),
    ("2026-03-31", "Example Platform Fee", "1400 Agent Rd", "Wilmington", "DE", "19801",
     "Credit card processing", 5.70, "B(1)", "cash", "", "Gross donations $1650; fee $5.70"),
    ("2026-02-14", "Example Venue", "1500 Event Blvd", "Phoenix", "AZ", "85003",
     "Fundraiser venue rental", 450.00, "B(1)", "cash", "", "Kickoff event 2/14"),
    ("2026-03-22", "Example Catering", "1600 Taco St", "Tempe", "AZ", "85281",
     "Volunteer food", 127.50, "B(1)", "cash", "", ""),
    ("2026-03-10", "Example Web Services", "1700 Nameserver Ln", "Scottsdale", "AZ", "85260",
     "Domain + email hosting", 25.99, "B(1)", "credit", "", "Annual renewal"),
]


def write_disbursements():
    cols = ["date", "payee", "address", "city", "state", "zip", "category",
            "amount", "schedule", "payment_method", "recipient_committee_id", "note"]
    rows = [dict(zip(cols, row)) for row in DISBURSEMENTS]
    xlsx = OUT / "sample_disbursements.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Disbursements"
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = PatternFill("solid", start_color="305496")
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(rows, start=2):
        for i, c in enumerate(cols, start=1):
            ws.cell(row=r, column=i, value=row[c])
    for col_idx, w in enumerate([12, 28, 28, 18, 8, 10, 30, 12, 10, 16, 24, 30], start=1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"
    wb.save(xlsx)
    print(f"wrote {xlsx}")

    # CSV version too
    csv_path = OUT / "sample_disbursements.csv"
    with open(csv_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols); w.writeheader()
        for row in rows:
            w.writerow(row)
    print(f"wrote {csv_path}")


# ---------------------------------------------------------------------------
# Debts (Schedule B(12))
# ---------------------------------------------------------------------------
DEBTS = [
    ("2026-03-19", "Example Unpaid Vendor", "1800 Debt Ln", "Phoenix", "AZ", "85016",
     "Print / mailer", 358.03, "2026-04-16", "Invoice dated 3/19; scheduled to pay 4/16"),
    ("2026-03-30", "Example Unpaid Vendor", "1800 Debt Ln", "Phoenix", "AZ", "85016",
     "Print / mailer", 551.86, "2026-04-16", "Invoice dated 3/30; scheduled to pay 4/16"),
]


def write_debts():
    cols = ["date_incurred", "creditor", "address", "city", "state", "zip",
            "type_of_debt", "amount_outstanding", "scheduled_payment_date", "note"]
    rows = [dict(zip(cols, row)) for row in DEBTS]
    xlsx = OUT / "sample_debts.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Debts"
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = PatternFill("solid", start_color="305496")
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(rows, start=2):
        for i, c in enumerate(cols, start=1):
            ws.cell(row=r, column=i, value=row[c])
    ws.freeze_panes = "A2"
    wb.save(xlsx)
    print(f"wrote {xlsx}")


# ---------------------------------------------------------------------------
# README for the samples directory
# ---------------------------------------------------------------------------
README_MD = """# Sample data for testing openfile-az

Everything in this directory is **100% synthetic** — no real donors, committees,
vendors, addresses, or filing IDs. Safe to share, post publicly, and commit.
These files exist so you can:

1. Drop them into the [web UI](https://azfile.energyfreedom.team) to see the tool work
   end-to-end without typing in your own data first.
2. Use them as starting templates for your own filing — copy, edit, replace
   all fields.
3. Run automated tests of the classifier and PDF generator against known
   inputs.

## Contents

| File | Purpose |
|------|---------|
| `sample_committee_config.yaml` | Two joint-fundraising committees |
| `sample_donors.xlsx` | 41 donors: 3 large in-state, 30 small in-state, 5 OOS, 2 PACs, 1 in-kind |
| `sample_donors.csv` | Same data, CSV form |
| `sample_donors_givebutter.csv` | Same individual donors, Givebutter export shape |
| `sample_donors_actblue.csv` | Same individual donors, ActBlue export shape (43 columns, "Donor X" prefixed, Contribution Datetime) — the most common platform export pattern |
| `sample_disbursements.xlsx` | 17 operating expenses (exceeds 5-row page → generates Excel supplement) |
| `sample_disbursements.csv` | Same data, CSV form |
| `sample_debts.xlsx` | 2 unpaid obligations (Schedule B(12)) |

## Regenerating

```bash
python3 web/scripts/generate_samples.py
```

Deterministic — same seed produces identical files every run.
"""


def write_samples_readme():
    out = OUT / "README.md"
    out.write_text(README_MD)
    print(f"wrote {out}")


# ---------------------------------------------------------------------------

def main():
    write_committee_config()
    rows = build_donor_rows()
    write_donors(rows)
    write_disbursements()
    write_debts()
    write_samples_readme()
    print(f"\nAll sample files in: {OUT}")


if __name__ == "__main__":
    main()
