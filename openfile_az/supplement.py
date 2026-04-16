"""Generate supplemental Excel sheets listing the FULL Schedule B(1) disbursements
for each Q1 PDF (since AZ form's page 40 only fits 5 rows).

Each Excel mirrors the AZ Schedule B(1) layout so it can be attached to the PDF
filing as official itemization for rows beyond the first 5.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from build_schedules import all_b1_rows, fmt

HERE = Path(__file__).parent
OUT_DIR = HERE.parent  # /Users/johnt/outputs

BOLD = Font(bold=True, name="Arial", size=10)
BOLD_WHITE = Font(bold=True, name="Arial", size=10, color="FFFFFF")
REG = Font(name="Arial", size=10)
ITAL = Font(name="Arial", size=9, italic=True, color="555555")
HEADER_FILL = PatternFill("solid", start_color="305496")
SUBTOTAL_FILL = PatternFill("solid", start_color="D9E1F2")
TOTAL_FILL = PatternFill("solid", start_color="BDD7EE")
PAGE40_FILL = PatternFill("solid", start_color="FFF2CC")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '$#,##0.00;($#,##0.00);"-"'


def build_supplement(committee, fname):
    is_john = committee == "John"
    rows = all_b1_rows(committee)
    n_pages = max(1, (len(rows) + 4) // 5)
    grand_total = sum(r["amount"] for r in rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule B(1) Supplement"

    # ---- Title block ----
    ws["A1"] = f"{committee} Election Committee — Schedule B(1) Supplement"
    ws["A1"].font = Font(bold=True, name="Arial", size=14)
    ws.merge_cells("A1:I1")
    ws["A2"] = "Q1 2026 Report (Jan 1 – Mar 31, 2026)  |  Disbursements for Operating Expenses"
    ws["A2"].font = Font(bold=True, name="Arial", size=11, color="305496")
    ws.merge_cells("A2:I2")
    ws["A3"] = (f"Attached to AZ SOS Committee Campaign Finance Report (filing date 04/15/2026). "
                f"This supplement lists ALL Schedule B(1) operating-expense disbursements for the period. "
                f"The first 5 highest-dollar rows also appear on page 40 of the main PDF. "
                f"This sheet is the authoritative itemization for rows 6–{len(rows)}.")
    ws["A3"].font = ITAL
    ws["A3"].alignment = LEFT
    ws.merge_cells("A3:I3")
    ws.row_dimensions[3].height = 42
    ws["A5"] = f"Schedule B(1), page 1 of {n_pages} (this supplement)"
    ws["A5"].font = ITAL
    ws.merge_cells("A5:I5")

    # ---- Column headers (mirror AZ form columns) ----
    headers = ["#", "Date", "Payee Name", "Street Address", "City", "State", "ZIP",
               "Type of Operating Expense Paid", "Amount ($)"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=7, column=i, value=h)
        c.font = BOLD_WHITE
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[7].height = 30

    # ---- Rows ----
    r_idx = 8
    for i, row in enumerate(rows, start=1):
        on_page40 = i <= 5
        cells = [i, row["date"], row["name"], row["address"], row["city"],
                 row["state"], row["zip"], row["category"], row["amount"]]
        for cc, v in enumerate(cells, start=1):
            c = ws.cell(row=r_idx, column=cc, value=v)
            c.font = REG
            c.alignment = LEFT if cc <= 8 else RIGHT
            c.border = BORDER
            if cc == 9:
                c.number_format = MONEY
            if on_page40:
                c.fill = PAGE40_FILL
        r_idx += 1

    # ---- Totals ----
    page40_subtotal = sum(r["amount"] for r in rows[:5])
    supplement_subtotal = sum(r["amount"] for r in rows[5:])

    r_idx += 1
    ws.cell(row=r_idx, column=8, value="On main PDF page 40 (rows 1–5):").font = BOLD
    ws.cell(row=r_idx, column=8).alignment = RIGHT
    c = ws.cell(row=r_idx, column=9, value=page40_subtotal)
    c.font = BOLD; c.fill = SUBTOTAL_FILL; c.number_format = MONEY; c.border = BORDER
    r_idx += 1
    ws.cell(row=r_idx, column=8, value=f"On this supplement (rows 6–{len(rows)}):").font = BOLD
    ws.cell(row=r_idx, column=8).alignment = RIGHT
    c = ws.cell(row=r_idx, column=9, value=supplement_subtotal)
    c.font = BOLD; c.fill = SUBTOTAL_FILL; c.number_format = MONEY; c.border = BORDER
    r_idx += 1
    ws.cell(row=r_idx, column=8, value="Schedule B(1) GRAND TOTAL (transfer to Summary of Disbursements, line 1):").font = BOLD
    ws.cell(row=r_idx, column=8).alignment = RIGHT
    c = ws.cell(row=r_idx, column=9, value=grand_total)
    c.font = BOLD; c.fill = TOTAL_FILL; c.number_format = MONEY; c.border = BORDER

    # ---- Notes ----
    r_idx += 2
    ws.cell(row=r_idx, column=1, value="Notes:").font = BOLD
    notes = [
        "• All disbursements paid by cash/credit card; no in-kind operating expenses for this period.",
        "• donation platforms platform fees are the campaign-borne portion (gross minus donor-covered).",
        "• 50/50 splits (Example Print Vendor A, Example Print Vendor B, etc.) reflect joint fundraising committee allocation per agreement with the other co-fundraising committee.",
        "• Yellow-shaded rows are also itemized on page 40 of the main PDF (the 5 highest-dollar entries).",
        "• Amounts here reconcile to: page 40 sumb1 (page subtotal) + this supplement subtotal = Schedule B(1) grand total = page 4 line 1 (Cash column).",
    ]
    for n in notes:
        r_idx += 1
        ws.cell(row=r_idx, column=1, value=n).font = REG
        ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=9)

    # ---- Column widths ----
    widths = {"A": 5, "B": 12, "C": 36, "D": 28, "E": 18, "F": 8, "G": 8, "H": 32, "I": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    out_path = OUT_DIR / fname
    wb.save(out_path)
    print(f"Wrote {out_path}  ({len(rows)} rows, grand total ${grand_total:,.2f})")


if __name__ == "__main__":
    build_supplement("John", "JaneExample_Q1_2026_AZ-SOS_ScheduleB1_Supplement.xlsx")
    build_supplement("Sara", "AlexExample_Q1_2026_AZ-SOS_ScheduleB1_Supplement.xlsx")
