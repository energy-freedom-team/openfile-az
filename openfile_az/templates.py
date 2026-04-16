"""Generate blank XLSX templates users fill out for donors / disbursements / debts.

Each template ships with:
  - A 'README' sheet explaining every column and allowed values
  - A 'Data' sheet with headers, formatting, example rows, and data validation
  - Column widths tuned for readability
"""
from __future__ import annotations
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

BOLD = Font(bold=True, name="Arial", size=10)
BOLD_WHITE = Font(bold=True, name="Arial", size=10, color="FFFFFF")
REG = Font(name="Arial", size=10)
ITAL = Font(name="Arial", size=9, italic=True, color="555555")
HEADER_FILL = PatternFill("solid", start_color="305496")
EXAMPLE_FILL = PatternFill("solid", start_color="FFF2CC")
REQ_FILL = PatternFill("solid", start_color="FCE4D6")  # required cols header tint
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '$#,##0.00;($#,##0.00);"-"'
DATE_FMT = "yyyy-mm-dd"


def _write_headers(ws, columns):
    for i, col in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i, value=col["name"])
        c.font = BOLD_WHITE
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[1].height = 28


def _write_examples(ws, columns, examples, start_row=2):
    for r, row in enumerate(examples, start=start_row):
        for ci, col in enumerate(columns, start=1):
            v = row.get(col["key"], "")
            c = ws.cell(row=r, column=ci, value=v)
            c.font = REG
            c.alignment = LEFT
            c.border = BORDER
            c.fill = EXAMPLE_FILL
            if col.get("fmt") == "money":
                c.number_format = MONEY
                c.alignment = RIGHT
            elif col.get("fmt") == "date":
                c.number_format = DATE_FMT


def _column_widths(ws, columns):
    for i, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = col.get("width", 14)


def _add_readme(wb, title, intro, columns, notes):
    ws = wb.create_sheet("README", 0)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, name="Arial", size=16)
    ws.merge_cells("A1:C1")
    ws["A2"] = intro
    ws["A2"].font = ITAL
    ws["A2"].alignment = LEFT
    ws.merge_cells("A2:C2")
    ws.row_dimensions[2].height = 52

    ws["A4"] = "Column"; ws["B4"] = "Required?"; ws["C4"] = "Description / allowed values"
    for c in (ws["A4"], ws["B4"], ws["C4"]):
        c.font = BOLD_WHITE; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[4].height = 24

    r = 5
    for col in columns:
        c1 = ws.cell(row=r, column=1, value=col["name"]); c1.font = BOLD; c1.border = BORDER
        c2 = ws.cell(row=r, column=2, value="Yes" if col.get("required") else "")
        c2.font = REG; c2.alignment = CENTER; c2.border = BORDER
        c3 = ws.cell(row=r, column=3, value=col.get("help", ""))
        c3.font = REG; c3.alignment = LEFT; c3.border = BORDER
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Notes:").font = BOLD
    for n in notes:
        r += 1
        ws.cell(row=r, column=1, value=n).font = REG
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 82


# ---------- Donors ----------
DONOR_COLS = [
    {"key": "date", "name": "date", "required": True, "width": 12, "fmt": "date",
     "help": "Date the contribution was received. Format YYYY-MM-DD."},
    {"key": "first", "name": "first", "required": False, "width": 14,
     "help": "Donor's first name. Leave blank for PACs / org donors (use 'last' for the org name)."},
    {"key": "last", "name": "last", "required": True, "width": 20,
     "help": "Donor's last name (individuals) or full legal name of PAC / org / committee."},
    {"key": "address", "name": "address", "required": True, "width": 28,
     "help": "Street address, suite/apt if any."},
    {"key": "city", "name": "city", "required": True, "width": 14,
     "help": "City."},
    {"key": "state", "name": "state", "required": True, "width": 8,
     "help": "Two-letter state code (AZ, CA, NY, DC, ...). Determines in-state vs out-of-state."},
    {"key": "zip", "name": "zip", "required": True, "width": 10,
     "help": "5-digit ZIP code."},
    {"key": "email", "name": "email", "required": False, "width": 22,
     "help": "Donor email (optional; used for matching across platforms, not reported)."},
    {"key": "occupation", "name": "occupation", "required": False, "width": 18,
     "help": "Individual donor occupation (optional; reported on Schedule A(1) when available)."},
    {"key": "employer", "name": "employer", "required": False, "width": 18,
     "help": "Individual donor employer (optional; reported on Schedule A(1) when available)."},
    {"key": "source", "name": "source", "required": True, "width": 22,
     "help": "Where the money came from. Use any stable label you choose — a fundraiser "
             "campaign name, a platform label, 'check', 'wire', 'bank', etc. For joint "
             "fundraising, this must match a 'campaign' defined in committee_config.yaml "
             "so the split rule applies automatically."},
    {"key": "amount", "name": "amount", "required": True, "width": 12, "fmt": "money",
     "help": "Gross contribution amount in USD. For joint fundraising, enter the GROSS before "
             "the committee split — the tool will divide it per config rules."},
    {"key": "kind", "name": "kind", "required": True, "width": 10,
     "help": "One of: cash | in_kind | loan. 'cash' includes check, card, ACH, wire."},
    {"key": "donor_type", "name": "donor_type", "required": True, "width": 16,
     "help": "One of: individual | pac | party | corp | llc | union | candidate | self. "
             "Determines schedule classification."},
    {"key": "pac_id", "name": "pac_id", "required": False, "width": 16,
     "help": "FEC committee ID (e.g., C00000000) or AZ committee ID. Required when donor_type "
             "is pac / party / candidate."},
    {"key": "note", "name": "note", "required": False, "width": 38,
     "help": "Free-form note — context, invoice refs, in-kind description, etc."},
]

DONOR_EXAMPLES = [
    {"date": "2026-02-20", "first": "Steven", "last": "Placeholder", "address": "123 E Chandler Blvd",
     "city": "Chandler", "state": "AZ", "zip": "85226", "email": "steve@example.com",
     "occupation": "Attorney", "employer": "ASU",
     "source": "Example Joint Campaign", "amount": 250.00,
     "kind": "cash", "donor_type": "individual", "pac_id": "",
     "note": "Joint joint fundraiser; 50/50 split handled by config"},
    {"date": "2026-03-24", "first": "", "last": "Example Climate PAC",
     "address": "100 Example Ave", "city": "Washington", "state": "DC",
     "zip": "20003", "email": "", "occupation": "", "employer": "",
     "source": "check", "amount": 2000.00, "kind": "cash",
     "donor_type": "pac", "pac_id": "C00000000", "note": "FEC PAC"},
    {"date": "2026-03-31", "first": "", "last": "Example Advocacy PAC",
     "address": "200 Example Blvd", "city": "Los Angeles", "state": "CA", "zip": "90029",
     "email": "", "occupation": "", "employer": "",
     "source": "in_kind_services", "amount": 159.70, "kind": "in_kind",
     "donor_type": "pac", "pac_id": "C00000001",
     "note": "March staff/phone/text services"},
]


def _build_donors_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb["Sheet"])
    _add_readme(
        wb,
        title="openfile-az — Donors template",
        intro=("Fill in one row per contribution. Headers in row 1 are required. Yellow rows "
               "below row 1 are EXAMPLES — delete them before generating your filing. Columns "
               "marked 'Required: Yes' must be populated."),
        columns=DONOR_COLS,
        notes=[
            "• The 'state' column drives in-state vs out-of-state classification (schedule A(1)(a)/(b) vs A(1)(c)).",
            "• Amount >$100 AND state=AZ AND donor_type=individual → Schedule A(1)(a); otherwise aggregated on A(1)(b).",
            "• donor_type=pac → Schedule A(1)(e); requires pac_id.",
            "• kind=in_kind → Schedule A(5) regardless of donor_type.",
            "• Joint-fundraising campaigns: enter the GROSS amount; config.yaml splits it.",
            "• Dates must be within the report's date range defined in committee_config.yaml.",
        ],
    )
    ws = wb.create_sheet("Data")
    _write_headers(ws, DONOR_COLS)
    _write_examples(ws, DONOR_COLS, DONOR_EXAMPLES)
    _column_widths(ws, DONOR_COLS)
    # Data validation: kind + donor_type
    kind_dv = DataValidation(type="list", formula1='"cash,in_kind,loan"', allow_blank=False)
    kind_dv.add("M2:M1000")
    ws.add_data_validation(kind_dv)
    dtype_dv = DataValidation(
        type="list",
        formula1='"individual,pac,party,corp,llc,union,candidate,self"',
        allow_blank=False,
    )
    dtype_dv.add("N2:N1000")
    ws.add_data_validation(dtype_dv)
    ws.freeze_panes = "A2"
    return wb


# ---------- Disbursements ----------
DISB_COLS = [
    {"key": "date", "name": "date", "required": True, "width": 12, "fmt": "date",
     "help": "Date payment was made (not the invoice date). YYYY-MM-DD."},
    {"key": "payee", "name": "payee", "required": True, "width": 30,
     "help": "Full legal name of payee (vendor, committee, individual)."},
    {"key": "address", "name": "address", "required": True, "width": 28,
     "help": "Payee street address."},
    {"key": "city", "name": "city", "required": True, "width": 14, "help": "Payee city."},
    {"key": "state", "name": "state", "required": True, "width": 8,
     "help": "Two-letter state code (or country for foreign vendors)."},
    {"key": "zip", "name": "zip", "required": True, "width": 10, "help": "ZIP / postal code."},
    {"key": "category", "name": "category", "required": True, "width": 32,
     "help": "Short description of the expense (what the form calls 'Type of Operating Expense "
             "Paid' for B(1))."},
    {"key": "amount", "name": "amount", "required": True, "width": 12, "fmt": "money",
     "help": "This committee's share of the payment in USD. For 50/50 shared expenses, enter "
             "only this committee's share (not the gross)."},
    {"key": "schedule", "name": "schedule", "required": True, "width": 16,
     "help": "Target AZ schedule: B(1) | B(2a) | B(2b) | B(2c) | B(6) | B(7) | B(10) | B(11). "
             "If left blank, the classifier defaults to B(1) Operating Expenses."},
    {"key": "payment_method", "name": "payment_method", "required": True, "width": 14,
     "help": "cash | credit. 'cash' covers check / ACH / wire / debit. 'credit' = credit-card "
             "charge paid later."},
    {"key": "recipient_committee_id", "name": "recipient_committee_id", "required": False,
     "width": 22, "help": "For schedule B(2a-c) / B(10): FEC or AZ ID of the receiving committee."},
    {"key": "note", "name": "note", "required": False, "width": 38,
     "help": "Free-form note — invoice number, 50/50 split context, etc."},
]

DISB_EXAMPLES = [
    {"date": "2026-03-31", "payee": "Example Print Vendor A", "address": "PO Box 1234",
     "city": "Phoenix", "state": "AZ", "zip": "85001",
     "category": "Print / mailer", "amount": 509.26, "schedule": "B(1)",
     "payment_method": "cash", "recipient_committee_id": "",
     "note": "50/50 with co-fundraising committee"},
    {"date": "2026-04-01", "payee": "Example GOTV Vendor",
     "address": "1500 Example Dr",
     "city": "Example City", "state": "CA", "zip": "92270",
     "category": "GOTV phone calls (inv 32011-1: 49-491 Final Push Live)",
     "amount": 666.66, "schedule": "B(1)", "payment_method": "cash",
     "recipient_committee_id": "", "note": ""},
]


def _build_disbursements_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb["Sheet"])
    _add_readme(
        wb,
        title="openfile-az — Disbursements template",
        intro=("One row per payment made by this committee. 50/50 shared expenses: enter THIS "
               "committee's share only. Use payment_method='credit' for unpaid credit-card "
               "charges that need to appear on Schedule B(12) Outstanding Debts instead."),
        columns=DISB_COLS,
        notes=[
            "• Schedule B(1) = Operating Expenses (most rows). Leave 'schedule' blank to default here.",
            "• Schedule B(10) = Joint Fundraising / Shared Expense Payments to another committee.",
            "• Schedule B(2a-c) = Monetary Contributions made to other candidate committees / PACs / parties.",
            "• If the committee has too many rows to fit the AZ form's 5-per-page limit, the tool auto-generates an Excel supplement for overflow rows.",
        ],
    )
    ws = wb.create_sheet("Data")
    _write_headers(ws, DISB_COLS)
    _write_examples(ws, DISB_COLS, DISB_EXAMPLES)
    _column_widths(ws, DISB_COLS)
    sched_dv = DataValidation(
        type="list",
        formula1='"B(1),B(2a),B(2b),B(2c),B(6),B(7),B(10),B(11)"',
        allow_blank=True,
    )
    sched_dv.add("I2:I1000")
    ws.add_data_validation(sched_dv)
    pm_dv = DataValidation(type="list", formula1='"cash,credit"', allow_blank=False)
    pm_dv.add("J2:J1000")
    ws.add_data_validation(pm_dv)
    ws.freeze_panes = "A2"
    return wb


# ---------- Debts ----------
DEBTS_COLS = [
    {"key": "date_incurred", "name": "date_incurred", "required": True, "width": 14, "fmt": "date",
     "help": "Date the debt was incurred (invoice date)."},
    {"key": "creditor", "name": "creditor", "required": True, "width": 28,
     "help": "Creditor's legal name (vendor that is owed money)."},
    {"key": "address", "name": "address", "required": True, "width": 28, "help": "Creditor address."},
    {"key": "city", "name": "city", "required": True, "width": 14},
    {"key": "state", "name": "state", "required": True, "width": 8},
    {"key": "zip", "name": "zip", "required": True, "width": 10},
    {"key": "type_of_debt", "name": "type_of_debt", "required": True, "width": 34,
     "help": "Description of the underlying expense (mirrors B(1) category)."},
    {"key": "amount_outstanding", "name": "amount_outstanding", "required": True,
     "width": 14, "fmt": "money",
     "help": "Amount still owed as of the report date. When partially paid, enter the remaining "
             "balance."},
    {"key": "scheduled_payment_date", "name": "scheduled_payment_date", "required": False,
     "width": 20, "fmt": "date",
     "help": "Expected payment date (informational)."},
    {"key": "note", "name": "note", "required": False, "width": 38},
]

DEBTS_EXAMPLES = [
    {"date_incurred": "2026-03-19", "creditor": "Example Unpaid Vendor",
     "address": "PO Box (TBD)", "city": "Phoenix", "state": "AZ", "zip": "85016",
     "type_of_debt": "Print / mailer (D6 mailer)",
     "amount_outstanding": 358.03, "scheduled_payment_date": "2026-04-16",
     "note": "Unpaid as of report date"},
]


def _build_debts_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb["Sheet"])
    _add_readme(
        wb,
        title="openfile-az — Outstanding Debts template (Schedule B(12))",
        intro=("One row per unpaid obligation outstanding as of the end of the report period. "
               "Debts are carried forward automatically to every subsequent report until paid."),
        columns=DEBTS_COLS,
        notes=[
            "• Schedule B(12) flows to Summary of Disbursements page 4 line 12 (Equity column).",
            "• Once a debt is paid, it should appear in the disbursements spreadsheet for that period and be removed (or zeroed out) here.",
        ],
    )
    ws = wb.create_sheet("Data")
    _write_headers(ws, DEBTS_COLS)
    _write_examples(ws, DEBTS_COLS, DEBTS_EXAMPLES)
    _column_widths(ws, DEBTS_COLS)
    ws.freeze_panes = "A2"
    return wb


# ---------- Public API ----------
def build_all(dest_dir: str | Path) -> list[Path]:
    """Write all three templates into dest_dir. Returns list of paths written."""
    d = Path(dest_dir)
    d.mkdir(parents=True, exist_ok=True)
    written = []
    for name, builder in [
        ("donors_template.xlsx", _build_donors_workbook),
        ("disbursements_template.xlsx", _build_disbursements_workbook),
        ("debts_template.xlsx", _build_debts_workbook),
    ]:
        p = d / name
        builder().save(p)
        written.append(p)
    return written


if __name__ == "__main__":
    import sys
    dest = sys.argv[1] if len(sys.argv) > 1 else "./templates"
    for p in build_all(dest):
        print(f"wrote {p}")
